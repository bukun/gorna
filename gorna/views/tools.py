import openpyxl


def load_excel(path):
    '''
    加载 Excel 中的数据。
    '''
    wb = openpyxl.load_workbook(path)
    worksheet = wb.active
    mcol = worksheet.max_column
    mrow = worksheet.max_row

    xlsx_title = worksheet.cell(row=1, column=2).value
    danwei_arr = ['']
    for idx in range(1, mcol + 1):
        danwei_arr.append(worksheet.cell(row=7, column=idx).value)

    code_arr = ['']
    for idx in range(1, mcol + 1):
        code_arr.append(worksheet.cell(row=8, column=idx).value)

    t_arr = ['ID']
    for idx in range(1, mcol + 1):
        t_arr.append(worksheet.cell(row=6, column=idx).value)

    d_arr = []

    id_index = 1
    for idx_row in range(10, mrow + 1):
        r_arr = [id_index]
        for idx_col in range(1, mcol + 1):
            r_arr.append(worksheet.cell(row=idx_row, column=idx_col).value)
        d_arr.append(r_arr)
        id_index = id_index + 1

    map_arr = {}
    data_stgart_m = code_arr.index('D')
    map_start_m = code_arr.index('CS') if 'CS' in code_arr else -1
    if map_start_m != -1:
        for idx_row in range(10, mrow + 1):
            map_arr[worksheet.cell(row=idx_row,
                                   column=map_start_m).value] = worksheet.cell(row=idx_row,
                                                                               column=data_stgart_m).value
    print('x' * 40)
    print(map_arr)

    totalArray = {
        'titles': t_arr,
        'values': d_arr,
        'codes': code_arr,
        'danwei': danwei_arr,
        'data_start_m': data_stgart_m,
        'xlsx_title': xlsx_title,
        'map_arr': map_arr
    }

    return totalArray
